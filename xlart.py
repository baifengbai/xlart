from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image


def resize_image(filename, new_filename, width, height):
    img = Image.open(filename)
    new_img = img.convert('RGB')
    new_img = new_img.resize((width, height))
    new_img.save(new_filename, 'JPEG', optimize=True)


def image_to_xlsx(image_file, xlsx_file):
    img = Image.open(image_file)
    pixel = img.load()
    wb = Workbook()
    ws = wb.active

    for i in range(img.width):
        for j in range(img.height):
            ws.cell(
                row=j + 1, column=i + 1).fill = PatternFill(
                    fgColor=Color('{0:02x}{1:02x}{2:02x}'.format(
                        pixel[i, j][0], pixel[i, j][1], pixel[i, j][2])),
                    fill_type='solid')

    for c in range(ws.max_column):
        ws.column_dimensions[get_column_letter(c + 1)].width = float(1.89)
    for r in range(ws.max_row):
        ws.row_dimensions[r + 1].height = float(14.4)

    wb.save(xlsx_file)
